import streamlit as st
import fitz
import google.generativeai as genai
import json
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
from datetime import datetime
import re
import os

# ── CONFIG —
GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "YOUR_GEMINI_API_KEY_HERE")

# ── Page setup ────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Invoxa — AI Invoice Extractor",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ── Global CSS ────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Syne:wght@400;600;700;800&family=DM+Sans:ital,wght@0,300;0,400;0,500;1,300&display=swap');

/* ── Reset & base ── */
*, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }

html, body, [data-testid="stAppViewContainer"] {
    background: #0a0a0f !important;
    color: #f0eee8;
    font-family: 'DM Sans', sans-serif;
}

[data-testid="stAppViewContainer"] > .main { background: #0a0a0f !important; }
[data-testid="stHeader"] { background: transparent !important; }
[data-testid="stSidebar"] { background: #0d0d14 !important; }
[data-testid="stToolbar"] { display: none; }
.block-container { padding: 0 !important; max-width: 100% !important; }
section[data-testid="stSidebar"] { display: none; }

/* ── Hide Streamlit chrome ── */
#MainMenu, footer, header { visibility: hidden; }

/* ── Hero section ── */
.hero {
    background: #0a0a0f;
    padding: 72px 60px 48px;
    position: relative;
    overflow: hidden;
    border-bottom: 1px solid rgba(255,255,255,0.06);
}
.hero::before {
    content: '';
    position: absolute;
    top: -120px; left: -80px;
    width: 520px; height: 520px;
    background: radial-gradient(circle, rgba(255,75,0,0.18) 0%, transparent 70%);
    pointer-events: none;
}
.hero::after {
    content: '';
    position: absolute;
    top: -60px; right: 80px;
    width: 380px; height: 380px;
    background: radial-gradient(circle, rgba(108,0,255,0.14) 0%, transparent 70%);
    pointer-events: none;
}
.hero-badge {
    display: inline-flex; align-items: center; gap: 8px;
    background: rgba(255,75,0,0.12);
    border: 1px solid rgba(255,75,0,0.35);
    border-radius: 99px;
    padding: 5px 14px;
    font-family: 'DM Sans', sans-serif;
    font-size: 12px; font-weight: 500;
    color: #ff7a45;
    letter-spacing: 0.04em;
    margin-bottom: 24px;
}
.hero-badge .dot {
    width: 6px; height: 6px;
    background: #ff4b00;
    border-radius: 50%;
    animation: pulse-dot 2s ease-in-out infinite;
}
@keyframes pulse-dot {
    0%, 100% { opacity: 1; transform: scale(1); }
    50% { opacity: 0.5; transform: scale(0.7); }
}
.hero h1 {
    font-family: 'Syne', sans-serif;
    font-size: clamp(42px, 6vw, 76px);
    font-weight: 800;
    line-height: 1.0;
    letter-spacing: -0.03em;
    color: #f5f2ec;
    margin-bottom: 20px;
}
.hero h1 .accent {
    background: linear-gradient(135deg, #ff4b00 0%, #ff9500 50%, #ffce00 100%);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    background-clip: text;
}
.hero p {
    font-family: 'DM Sans', sans-serif;
    font-size: 18px;
    font-weight: 300;
    color: rgba(240,238,232,0.55);
    max-width: 520px;
    line-height: 1.65;
    margin-bottom: 36px;
}
.hero-stats {
    display: flex; gap: 40px; flex-wrap: wrap;
}
.hero-stat {
    display: flex; flex-direction: column;
}
.hero-stat .num {
    font-family: 'Syne', sans-serif;
    font-size: 28px; font-weight: 700;
    background: linear-gradient(135deg, #ff4b00, #ffce00);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    background-clip: text;
}
.hero-stat .lbl {
    font-size: 12px; font-weight: 400;
    color: rgba(240,238,232,0.4);
    letter-spacing: 0.05em;
    text-transform: uppercase;
}

/* ── Upload zone ── */
.upload-section {
    padding: 48px 60px;
    background: #0a0a0f;
}
.upload-label {
    font-family: 'Syne', sans-serif;
    font-size: 13px; font-weight: 600;
    letter-spacing: 0.1em;
    text-transform: uppercase;
    color: rgba(240,238,232,0.4);
    margin-bottom: 16px;
}

[data-testid="stFileUploader"] {
    background: rgba(255,255,255,0.02) !important;
    border: 1.5px dashed rgba(255,75,0,0.3) !important;
    border-radius: 16px !important;
    padding: 40px !important;
    transition: border-color 0.2s, background 0.2s;
}
[data-testid="stFileUploader"]:hover {
    border-color: rgba(255,75,0,0.6) !important;
    background: rgba(255,75,0,0.03) !important;
}
[data-testid="stFileUploader"] label {
    color: rgba(240,238,232,0.5) !important;
    font-family: 'DM Sans', sans-serif !important;
}
[data-testid="stFileUploader"] small {
    color: rgba(240,238,232,0.3) !important;
}
[data-testid="stFileUploaderDropzoneInstructions"] div span {
    color: rgba(240,238,232,0.5) !important;
    font-family: 'DM Sans', sans-serif;
}

/* ── Stats row ── */
.stats-row {
    display: grid;
    grid-template-columns: repeat(4, 1fr);
    gap: 16px;
    margin: 32px 0 0;
}
.stat-card {
    background: rgba(255,255,255,0.03);
    border: 1px solid rgba(255,255,255,0.07);
    border-radius: 14px;
    padding: 20px 22px;
    position: relative;
    overflow: hidden;
}
.stat-card::before {
    content: '';
    position: absolute;
    top: 0; left: 0; right: 0;
    height: 2px;
    background: linear-gradient(90deg, #ff4b00, #ffce00);
}
.stat-card .s-num {
    font-family: 'Syne', sans-serif;
    font-size: 32px; font-weight: 700;
    color: #f5f2ec;
    line-height: 1;
}
.stat-card .s-lbl {
    font-size: 11px; font-weight: 500;
    color: rgba(240,238,232,0.35);
    text-transform: uppercase;
    letter-spacing: 0.08em;
    margin-top: 6px;
}

/* ── Extract button ── */
.stButton > button {
    background: linear-gradient(135deg, #ff4b00 0%, #ff7a00 100%) !important;
    color: #fff !important;
    border: none !important;
    border-radius: 12px !important;
    padding: 16px 36px !important;
    font-family: 'Syne', sans-serif !important;
    font-size: 15px !important;
    font-weight: 700 !important;
    letter-spacing: 0.02em !important;
    width: 100% !important;
    transition: all 0.2s !important;
    box-shadow: 0 4px 24px rgba(255,75,0,0.3) !important;
    margin-top: 24px !important;
}
.stButton > button:hover {
    transform: translateY(-2px) !important;
    box-shadow: 0 8px 32px rgba(255,75,0,0.45) !important;
}
.stButton > button:active { transform: translateY(0) !important; }

/* ── Warning / info ── */
[data-testid="stAlert"] {
    background: rgba(255,75,0,0.08) !important;
    border: 1px solid rgba(255,75,0,0.25) !important;
    border-radius: 10px !important;
    color: #ff9060 !important;
}

/* ── Progress ── */
[data-testid="stProgressBar"] > div > div {
    background: linear-gradient(90deg, #ff4b00, #ffce00) !important;
}

/* ── Results section ── */
.results-header {
    padding: 32px 60px 16px;
    border-top: 1px solid rgba(255,255,255,0.06);
}
.results-title {
    font-family: 'Syne', sans-serif;
    font-size: 22px; font-weight: 700;
    color: #f5f2ec;
}
.results-sub {
    font-size: 13px;
    color: rgba(240,238,232,0.35);
    margin-top: 4px;
}

/* ── Invoice result cards ── */
.inv-card {
    background: rgba(255,255,255,0.03);
    border: 1px solid rgba(255,255,255,0.07);
    border-radius: 16px;
    padding: 24px 28px;
    margin: 0 60px 16px;
    transition: border-color 0.2s;
}
.inv-card:hover { border-color: rgba(255,75,0,0.3); }
.inv-card-header {
    display: flex;
    justify-content: space-between;
    align-items: flex-start;
    margin-bottom: 20px;
    flex-wrap: wrap;
    gap: 12px;
}
.inv-card-title {
    font-family: 'Syne', sans-serif;
    font-size: 15px; font-weight: 700;
    color: #f5f2ec;
}
.inv-card-sub {
    font-size: 12px;
    color: rgba(240,238,232,0.35);
    margin-top: 3px;
}
.inv-total {
    font-family: 'Syne', sans-serif;
    font-size: 22px; font-weight: 800;
    background: linear-gradient(135deg, #ff4b00, #ffce00);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    background-clip: text;
}
.badge-ok {
    display: inline-block;
    background: rgba(0,210,110,0.12);
    border: 1px solid rgba(0,210,110,0.3);
    color: #00d26a;
    border-radius: 99px;
    padding: 3px 12px;
    font-size: 11px; font-weight: 600;
    letter-spacing: 0.05em;
    text-transform: uppercase;
}
.badge-err {
    display: inline-block;
    background: rgba(255,75,0,0.12);
    border: 1px solid rgba(255,75,0,0.3);
    color: #ff6b40;
    border-radius: 99px;
    padding: 3px 12px;
    font-size: 11px; font-weight: 600;
    letter-spacing: 0.05em;
    text-transform: uppercase;
}
.field-grid {
    display: grid;
    grid-template-columns: repeat(3, 1fr);
    gap: 16px;
    margin-bottom: 20px;
}
.field-item {}
.field-lbl {
    font-size: 10px; font-weight: 500;
    color: rgba(240,238,232,0.3);
    text-transform: uppercase;
    letter-spacing: 0.1em;
    margin-bottom: 4px;
}
.field-val {
    font-size: 14px; font-weight: 400;
    color: #f0eee8;
}

/* ── Expander ── */
[data-testid="stExpander"] {
    background: rgba(255,255,255,0.02) !important;
    border: 1px solid rgba(255,255,255,0.07) !important;
    border-radius: 14px !important;
    margin: 0 60px 14px !important;
}
[data-testid="stExpander"] summary {
    color: #f0eee8 !important;
    font-family: 'DM Sans', sans-serif !important;
    padding: 16px 20px !important;
}
[data-testid="stExpander"] summary:hover {
    background: rgba(255,75,0,0.04) !important;
    border-radius: 14px !important;
}

/* ── Dataframe ── */
[data-testid="stDataFrame"] {
    border: 1px solid rgba(255,255,255,0.07) !important;
    border-radius: 10px !important;
    overflow: hidden;
}

/* ── Download button ── */
.download-zone {
    padding: 32px 60px 60px;
}
[data-testid="stDownloadButton"] button {
    background: rgba(255,255,255,0.04) !important;
    border: 1.5px solid rgba(255,75,0,0.4) !important;
    color: #ff7a45 !important;
    border-radius: 12px !important;
    padding: 14px 28px !important;
    font-family: 'Syne', sans-serif !important;
    font-size: 14px !important;
    font-weight: 700 !important;
    width: 100% !important;
    transition: all 0.2s !important;
}
[data-testid="stDownloadButton"] button:hover {
    background: rgba(255,75,0,0.1) !important;
    border-color: rgba(255,75,0,0.7) !important;
    transform: translateY(-1px) !important;
}

/* ── Footer ── */
.footer {
    padding: 24px 60px;
    border-top: 1px solid rgba(255,255,255,0.05);
    display: flex;
    justify-content: space-between;
    align-items: center;
    flex-wrap: wrap;
    gap: 12px;
}
.footer-brand {
    font-family: 'Syne', sans-serif;
    font-size: 14px; font-weight: 700;
    color: rgba(240,238,232,0.3);
}
.footer-brand span {
    background: linear-gradient(135deg, #ff4b00, #ffce00);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    background-clip: text;
}
.footer-note {
    font-size: 12px;
    color: rgba(240,238,232,0.2);
}

/* ── Scrollbar ── */
::-webkit-scrollbar { width: 6px; }
::-webkit-scrollbar-track { background: #0a0a0f; }
::-webkit-scrollbar-thumb { background: rgba(255,75,0,0.3); border-radius: 3px; }
::-webkit-scrollbar-thumb:hover { background: rgba(255,75,0,0.5); }

/* ── Streamlit padding fix ── */
.element-container { padding: 0 !important; }
div[data-testid="column"] { padding: 4px 8px !important; }

</style>
""", unsafe_allow_html=True)


# ── Helpers ───────────────────────────────────────────────────────────────────

def extract_text_from_pdf(file_bytes: bytes) -> str:
    doc = fitz.open(stream=file_bytes, filetype="pdf")
    text = "".join(page.get_text() for page in doc)
    doc.close()
    return text.strip()


def extract_invoice_data(file_bytes: bytes, filename: str) -> dict:
    genai.configure(api_key=GEMINI_API_KEY)
    model = genai.GenerativeModel("gemini-2.5-flash")
    ext = filename.lower().split(".")[-1]
    is_image = ext in ("jpg", "jpeg", "png", "webp")

    PROMPT = """You are an invoice data extraction specialist.
Extract ALL fields from this invoice and return ONLY a valid JSON object — no explanation, no markdown, no backticks.

Required JSON structure:
{
  "vendor_name": "string",
  "vendor_address": "string or null",
  "vendor_email": "string or null",
  "vendor_phone": "string or null",
  "invoice_number": "string",
  "invoice_date": "string (DD/MM/YYYY if possible)",
  "due_date": "string or null",
  "bill_to": "string",
  "bill_to_address": "string or null",
  "subtotal": number or null,
  "tax_rate": "string or null",
  "tax_amount": number or null,
  "discount": number or null,
  "total_amount": number,
  "currency": "string (e.g. PKR, USD)",
  "payment_terms": "string or null",
  "notes": "string or null",
  "line_items": [
    {"description": "string", "quantity": number, "unit_price": number, "amount": number}
  ]
}
If a field is not found, use null. Numbers must be actual numbers with no commas or currency symbols."""

    try:
        if is_image:
            mt = {"jpg":"image/jpeg","jpeg":"image/jpeg","png":"image/png","webp":"image/webp"}.get(ext,"image/jpeg")
            response = model.generate_content([PROMPT, {"mime_type": mt, "data": file_bytes}])
        else:
            raw_text = extract_text_from_pdf(file_bytes)
            if not raw_text:
                return {"_filename": filename, "_status": "error",
                        "error": "No text found in PDF. Try uploading as a JPG image."}
            response = model.generate_content(f"{PROMPT}\n\nInvoice text:\n{raw_text}")

        raw = re.sub(r"^```(?:json)?\s*", "", response.text.strip())
        raw = re.sub(r"\s*```$", "", raw)
        data = json.loads(raw)
        data["_filename"] = filename
        data["_status"] = "success"
        return data

    except json.JSONDecodeError as e:
        return {"_filename": filename, "_status": "error", "error": f"Parse error: {e}"}
    except Exception as e:
        msg = str(e)
        if "api_key" in msg.lower() or "api key" in msg.lower():
            msg = "API key error — check your GEMINI_API_KEY in secrets."
        return {"_filename": filename, "_status": "error", "error": msg}


def build_excel(results: list) -> bytes:
    wb = openpyxl.Workbook()
    DARK = "0D0D14"; ORANGE = "FF4B00"; ORANGE2 = "FF9500"
    LIGHT_ORANGE = "FFF0E8"; WHITE = "FFFFFF"; GRAY = "F8F6F2"
    GREEN_B = "E8F8F0"; GREEN_T = "0A6640"; RED_B = "FEE8E8"; RED_T = "991B1B"

    def F(hex_c): return PatternFill("solid", fgColor=hex_c)
    def font(bold=False, color="1A1A2E", size=10, italic=False):
        return Font(name="Calibri", bold=bold, color=color, size=size, italic=italic)
    def border(b="all"):
        s = Side(style="thin", color="E8E4DE"); n = Side(style=None)
        return Border(left=s,right=s,top=s,bottom=s) if b=="all" else Border(bottom=s,left=n,right=n,top=n)
    def cw(ws, col, w): ws.column_dimensions[get_column_letter(col)].width = w

    success = [r for r in results if r.get("_status") == "success"]

    # ── Sheet 1: Summary ──
    ws = wb.active; ws.title = "Summary"; ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:K1")
    ws["A1"] = "⚡ INVOXA — Invoice Extraction Report"
    ws["A1"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
    ws["A1"].fill = F(DARK)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[1].height = 38

    ws.merge_cells("A2:K2")
    ws["A2"] = f"Generated {datetime.now().strftime('%d %B %Y at %H:%M')}  ·  {len(success)} invoices processed  ·  Powered by Gemini AI"
    ws["A2"].font = Font(name="Calibri", color="888888", size=9, italic=True)
    ws["A2"].fill = F("1A1A2E")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 6

    hdrs = ["#","File","Vendor","Invoice No.","Date","Bill To","Subtotal","Tax","Total","Currency","Status"]
    ws_  = [4,  26,    22,      15,            12,    22,       12,        10,   12,     10,         10]
    for c,(h,w) in enumerate(zip(hdrs,ws_),1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=9)
        cell.fill = F(ORANGE); cell.border = border()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cw(ws, c, w)
    ws.row_dimensions[4].height = 22

    for i, inv in enumerate(results, 1):
        ok = inv.get("_status") == "success"
        bg = WHITE if i%2 else GRAY
        vals = [i, inv.get("_filename",""),
                inv.get("vendor_name","—") if ok else "—",
                inv.get("invoice_number","—") if ok else "—",
                inv.get("invoice_date","—") if ok else "—",
                inv.get("bill_to","—") if ok else "—",
                inv.get("subtotal") if ok else None,
                inv.get("tax_amount") if ok else None,
                inv.get("total_amount") if ok else None,
                inv.get("currency","—") if ok else "—",
                "Success" if ok else "Error"]
        for c,v in enumerate(vals,1):
            cell = ws.cell(row=4+i, column=c, value=v)
            cell.font = font(size=9); cell.fill = F(bg); cell.border = border("bottom")
            cc = c in (1,5,7,8,9,10,11)
            cell.alignment = Alignment(horizontal="center" if cc else "left", vertical="center", indent=0 if cc else 1)
            if c in (7,8,9) and v is not None: cell.number_format = "#,##0.00"
            if c == 11:
                cell.font = Font(name="Calibri",bold=True,size=9,color=GREEN_T if ok else RED_T)
                cell.fill = F(GREEN_B if ok else RED_B)
                cell.alignment = Alignment(horizontal="center",vertical="center")
        ws.row_dimensions[4+i].height = 18

    if success:
        tr = 5 + len(results)
        grand = sum(float(r.get("total_amount") or 0) for r in success)
        ws.cell(tr,9,"Grand Total →").font = Font(name="Calibri",bold=True,size=9,color=ORANGE)
        ws.cell(tr,9).alignment = Alignment(horizontal="right")
        tc = ws.cell(tr,10,grand); tc.number_format="#,##0.00"
        tc.font = Font(name="Calibri",bold=True,size=11,color=ORANGE)
        tc.fill = F(LIGHT_ORANGE); tc.alignment = Alignment(horizontal="center")
        ws.row_dimensions[tr].height = 24
    ws.freeze_panes = "A5"

    # ── Sheet 2: Line Items ──
    ws2 = wb.create_sheet("Line Items"); ws2.sheet_view.showGridLines = False
    ws2.merge_cells("A1:G1"); ws2["A1"] = "Line Items Detail"
    ws2["A1"].font = Font(name="Calibri",bold=True,color="FFFFFF",size=13)
    ws2["A1"].fill = F(DARK); ws2["A1"].alignment = Alignment(horizontal="left",vertical="center",indent=2)
    ws2.row_dimensions[1].height = 32; ws2.row_dimensions[2].height = 6
    lh = ["File","Vendor","Invoice No.","Description","Qty","Unit Price","Amount"]
    lw = [26,22,15,36,7,13,13]
    for c,(h,w) in enumerate(zip(lh,lw),1):
        cell = ws2.cell(row=3,column=c,value=h)
        cell.font = Font(name="Calibri",bold=True,color="FFFFFF",size=9)
        cell.fill = F(ORANGE); cell.border = border()
        cell.alignment = Alignment(horizontal="center",vertical="center")
        ws2.column_dimensions[get_column_letter(c)].width = w
    ws2.row_dimensions[3].height = 22
    lr = 4; alt = True
    for inv in success:
        for item in (inv.get("line_items") or []):
            bg = WHITE if alt else GRAY
            rv = [inv.get("_filename",""),inv.get("vendor_name",""),inv.get("invoice_number",""),
                  item.get("description",""),item.get("quantity"),item.get("unit_price"),item.get("amount")]
            for c,v in enumerate(rv,1):
                cell = ws2.cell(row=lr,column=c,value=v)
                cell.font = font(size=9); cell.fill = F(bg); cell.border = border("bottom")
                cell.alignment = Alignment(horizontal="center" if c in (5,6,7) else "left",vertical="center",indent=0 if c in (5,6,7) else 1)
                if c in (6,7) and v is not None: cell.number_format = "#,##0.00"
            ws2.row_dimensions[lr].height = 17; lr+=1; alt=not alt
    ws2.freeze_panes = "A4"

    # ── Sheet 3: Full Details ──
    ws3 = wb.create_sheet("Full Details"); ws3.sheet_view.showGridLines = False
    ws3.column_dimensions["A"].width = 22; ws3.column_dimensions["B"].width = 38
    ws3.merge_cells("A1:B1"); ws3["A1"] = "Full Invoice Details"
    ws3["A1"].font = Font(name="Calibri",bold=True,color="FFFFFF",size=13)
    ws3["A1"].fill = F(DARK); ws3["A1"].alignment = Alignment(horizontal="left",vertical="center",indent=2)
    ws3.row_dimensions[1].height = 32
    fdr = 3
    fields = [("Vendor Name","vendor_name"),("Vendor Address","vendor_address"),
              ("Vendor Email","vendor_email"),("Vendor Phone","vendor_phone"),
              ("Invoice Number","invoice_number"),("Invoice Date","invoice_date"),
              ("Due Date","due_date"),("Bill To","bill_to"),("Bill To Address","bill_to_address"),
              ("Subtotal","subtotal"),("Tax Rate","tax_rate"),("Tax Amount","tax_amount"),
              ("Discount","discount"),("Total Amount","total_amount"),("Currency","currency"),
              ("Payment Terms","payment_terms"),("Notes","notes")]
    for inv in success:
        ws3.merge_cells(f"A{fdr}:B{fdr}"); ws3[f"A{fdr}"] = f"  {inv.get('_filename','')}"
        ws3[f"A{fdr}"].font = Font(name="Calibri",bold=True,color="FFFFFF",size=10)
        ws3[f"A{fdr}"].fill = F(ORANGE); ws3[f"A{fdr}"].alignment = Alignment(horizontal="left",vertical="center")
        ws3.row_dimensions[fdr].height = 22; fdr+=1
        for lbl,key in fields:
            val = inv.get(key); val = "—" if val is None else val
            lc = ws3.cell(fdr,1,lbl); lc.font = font(bold=True,size=9,color="555555")
            lc.fill = F(GRAY); lc.border = border("bottom"); lc.alignment = Alignment(horizontal="left",vertical="center",indent=2)
            vc = ws3.cell(fdr,2,val); vc.font = font(size=9); vc.fill = F(WHITE)
            vc.border = border("bottom"); vc.alignment = Alignment(horizontal="left",vertical="center",indent=2)
            if key in ("subtotal","tax_amount","discount","total_amount") and isinstance(val,(int,float)):
                vc.number_format = "#,##0.00"
            ws3.row_dimensions[fdr].height = 17; fdr+=1
        fdr+=2

    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# UI RENDER
# ══════════════════════════════════════════════════════════════════════════════

# ── Hero ──────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="hero">
    <div class="hero-badge"><span class="dot"></span>AI-Powered · Free to Use · No Sign-up</div>
    <h1>Turn invoices into<br><span class="accent">clean data</span><br>in seconds.</h1>
    <p>Drop PDF or image invoices. Invoxa reads every field automatically and exports a formatted Excel report — vendor, totals, line items, all of it.</p>
    <div class="hero-stats">
        <div class="hero-stat"><span class="num">1,500</span><span class="lbl">Free requests/day</span></div>
        <div class="hero-stat"><span class="num">&lt; 5s</span><span class="lbl">Per invoice</span></div>
        <div class="hero-stat"><span class="num">3</span><span class="lbl">Excel sheets</span></div>
        <div class="hero-stat"><span class="num">17+</span><span class="lbl">Fields extracted</span></div>
    </div>
</div>
""", unsafe_allow_html=True)

# ── Upload ─────────────────────────────────────────────────────────────────────
st.markdown('<div class="upload-section">', unsafe_allow_html=True)
st.markdown('<div class="upload-label">Upload your invoices</div>', unsafe_allow_html=True)

uploaded_files = st.file_uploader(
    "invoices",
    type=["pdf", "jpg", "jpeg", "png", "webp"],
    accept_multiple_files=True,
    label_visibility="collapsed",
)

if uploaded_files:
    pdf_c = sum(1 for f in uploaded_files if f.name.lower().endswith(".pdf"))
    img_c = len(uploaded_files) - pdf_c
    est   = len(uploaded_files) * 5

    st.markdown(f"""
    <div class="stats-row">
        <div class="stat-card"><div class="s-num">{len(uploaded_files)}</div><div class="s-lbl">Total files</div></div>
        <div class="stat-card"><div class="s-num">{pdf_c}</div><div class="s-lbl">PDFs</div></div>
        <div class="stat-card"><div class="s-num">{img_c}</div><div class="s-lbl">Images</div></div>
        <div class="stat-card"><div class="s-num">~{est}s</div><div class="s-lbl">Est. time</div></div>
    </div>
    """, unsafe_allow_html=True)

    if GEMINI_API_KEY == "YOUR_GEMINI_API_KEY_HERE":
        st.warning("⚠️ API key not configured. Add GEMINI_API_KEY to your Streamlit secrets.")
    else:
        if st.button("⚡ Extract All Invoices"):
            results = []
            prog  = st.progress(0, text="Initialising...")
            msg   = st.empty()
            for i, f in enumerate(uploaded_files):
                msg.info(f"Reading **{f.name}** — {i+1} of {len(uploaded_files)}")
                results.append(extract_invoice_data(f.read(), f.name))
                prog.progress((i+1)/len(uploaded_files), text=f"{i+1}/{len(uploaded_files)} complete")
            msg.empty(); prog.empty()
            st.session_state["results"] = results

st.markdown('</div>', unsafe_allow_html=True)

# ── Results ───────────────────────────────────────────────────────────────────
if "results" in st.session_state:
    results = st.session_state["results"]
    ok  = [r for r in results if r.get("_status") == "success"]
    err = [r for r in results if r.get("_status") != "success"]

    st.markdown(f"""
    <div class="results-header">
        <div class="results-title">Extraction complete</div>
        <div class="results-sub">{len(ok)} succeeded · {len(err)} failed · {len(ok)} invoices ready for download</div>
    </div>
    """, unsafe_allow_html=True)

    for inv in ok:
        total     = inv.get("total_amount")
        currency  = inv.get("currency") or ""
        total_str = f"{currency} {total:,.2f}" if isinstance(total, (int, float)) else "—"
        vendor    = inv.get("vendor_name") or "Unknown Vendor"
        inv_no    = inv.get("invoice_number") or "—"
        inv_date  = inv.get("invoice_date") or "—"

        with st.expander(f"✅  {inv.get('_filename','')}", expanded=False):
            st.markdown(f"""
            <div style="display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:20px;flex-wrap:wrap;gap:12px;">
                <div>
                    <div style="font-family:'Syne',sans-serif;font-size:17px;font-weight:700;color:#f5f2ec;">{vendor}</div>
                    <div style="font-size:12px;color:rgba(240,238,232,0.4);margin-top:3px;">Invoice {inv_no} · {inv_date}</div>
                </div>
                <div style="text-align:right;">
                    <div style="font-size:10px;text-transform:uppercase;letter-spacing:0.1em;color:rgba(240,238,232,0.3);margin-bottom:4px;">Total</div>
                    <div style="font-family:'Syne',sans-serif;font-size:24px;font-weight:800;background:linear-gradient(135deg,#ff4b00,#ffce00);-webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text;">{total_str}</div>
                </div>
            </div>
            <div class="field-grid">
            """, unsafe_allow_html=True)

            field_pairs = [
                ("Bill To",       inv.get("bill_to")),
                ("Due Date",      inv.get("due_date")),
                ("Subtotal",      f"{inv.get('subtotal'):,.2f}" if isinstance(inv.get('subtotal'),(int,float)) else inv.get('subtotal')),
                ("Tax Rate",      inv.get("tax_rate")),
                ("Tax Amount",    f"{inv.get('tax_amount'):,.2f}" if isinstance(inv.get('tax_amount'),(int,float)) else inv.get('tax_amount')),
                ("Payment Terms", inv.get("payment_terms")),
            ]
            cols = st.columns(3)
            for j, (lbl, val) in enumerate(field_pairs):
                with cols[j % 3]:
                    st.markdown(f"""
                    <div class="field-lbl">{lbl}</div>
                    <div class="field-val">{val or '—'}</div>
                    """, unsafe_allow_html=True)
                    st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)

            items = inv.get("line_items") or []
            if items:
                st.markdown("<div style='margin-top:12px;margin-bottom:8px;font-size:11px;text-transform:uppercase;letter-spacing:0.1em;color:rgba(240,238,232,0.3);'>Line Items</div>", unsafe_allow_html=True)
                df = pd.DataFrame(items)
                st.dataframe(df, use_container_width=True, hide_index=True)

    for inv in err:
        st.markdown(f"""
        <div style="margin:0 60px 12px;background:rgba(255,75,0,0.07);border:1px solid rgba(255,75,0,0.25);
            border-radius:12px;padding:16px 20px;display:flex;align-items:center;gap:12px;">
            <span class="badge-err">Failed</span>
            <span style="font-size:13px;color:rgba(240,238,232,0.6);">{inv.get('_filename','')} — {inv.get('error','Unknown error')}</span>
        </div>
        """, unsafe_allow_html=True)

    # ── Download ──
    if ok:
        st.markdown('<div class="download-zone">', unsafe_allow_html=True)
        excel_bytes = build_excel(results)
        fname = f"invoxa_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        st.download_button(
            label=f"⬇️  Download Excel Report  ({len(ok)} invoice{'s' if len(ok)>1 else ''})",
            data=excel_bytes,
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
        st.markdown("""
        <div style="text-align:center;margin-top:12px;font-size:12px;color:rgba(240,238,232,0.25);">
            3 sheets included: Summary · Line Items · Full Details
        </div>
        """, unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)

# ── Footer ────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="footer">
    <div class="footer-brand"><span>Invoxa</span> — AI Invoice Extractor</div>
    <div class="footer-note">Powered by Gemini 1.5 Flash · Built with Streamlit · Free to use</div>
</div>
""", unsafe_allow_html=True)
